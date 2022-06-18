import PyPDF2
from PyPDF2 import PdfFileReader
import os
import re
import openpyxl

excel_document = openpyxl.load_workbook('biofire_pdf_data.xlsx')
excel_sheet = excel_document['Sheet1']

#loop on files and extract what we need from PDFS
for file_name in os.listdir('C:/Users/wwwya/Desktop/pdf/extract_text'): #loop on files
    print(file_name)
    load_pdf = open(r'C:\\Users\wwwya\Desktop\pdf\\extract_text\\' + file_name,'rb')
    read_pdf = PyPDF2.PdfFileReader(load_pdf)                           #load entire pdf in variable
    page_count = read_pdf.getNumPages()
    first_page = read_pdf.getPage(0)                                    #read only first page
    page_content = first_page.extractText()                             #extract sting output
    page_content = page_content.replace('\n','')                        #replaces new lines in text
    print(page_content)
    print("--------------------------------------------")

    #Sample ID extraction
    sample_id = re.search(r'ID\:(.*?)Run', page_content).group(1)
    print(sample_id)

    #Date extraction
    sample_date = re.search(r'Date\:(.*?)\s', page_content).group(1)
    print(sample_date)

    #Time Extraction
    sample_time = re.search(r'[2022-3000]\s(.*?)Detected', page_content).group(1)
    print(sample_time)

    #Positives extraction
    sample_positive = re.search(r'Detected\:(.*?)Controls',page_content).group(1)

    result_covid = sample_positive.find("Coronavirus2")
    if result_covid == -1:
        result_covid = ("Not Detected")
    else:
        result_covid = ("Detected")
        print("Covid positive")

    result_fluA = sample_positive.find("H1-2009") or sample_positive.find("InfluenzaAH3")
    if result_fluA == -1:
        result_fluA = ("Not Detected")
    else:
        result_fluA = ("Detected")
        print("FluA Positive")

    result_fluB = sample_positive.find("InfluenzaB")
    if result_fluB == -1:
        result_fluB = ("Not Detected")
    else:
        result_fluB = ("Detected")
        print("FluB Positive")

    result_rsv = sample_positive.find("RespiratorySyncytialVirus")
    if result_rsv == -1:
        result_rsv = ("Not Detected")
    else:
        result_rsv = ("Detected")
        print("RSV Positive")

    print(sample_positive)

    #Technician ID extraction
    sample_technician = re.findall(r'\((.*?)\)',page_content)
    for x in sample_technician:
        x = sample_technician[-1]
    print(sample_technician[-1])

    #Transfer the data to excel

    #Tell excel the max collums you will use
    last_row_number = excel_sheet.max_row
    print(last_row_number)

    #Tell python to fill up excel

    excel_sheet.cell(column = 1, row = last_row_number + 1).value = sample_id
    excel_sheet.cell(column = 2, row = last_row_number + 1).value = x
    excel_sheet.cell(column = 3, row = last_row_number + 1).value = sample_date
    excel_sheet.cell(column = 4, row = last_row_number + 1).value = sample_time
    excel_sheet.cell(column = 5, row = last_row_number + 1).value = result_covid
    excel_sheet.cell(column = 6, row = last_row_number + 1).value = result_fluA
    excel_sheet.cell(column = 7, row = last_row_number + 1).value = result_fluB
    excel_sheet.cell(column = 8, row = last_row_number + 1).value = result_rsv

    #This saves the file

    #excel_document.save('biofire_pdf_data.xlsx')